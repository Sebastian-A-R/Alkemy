import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from unidecode import unidecode


if __name__ == "__main__":


    a = pd.read_csv("cine.csv")
    b = pd.read_csv("museo.csv")
    c = pd.read_csv("biblioteca_popular.csv")
    columnas_a_filtrar=["cod_loc","idprovincia","iddepartamento","categoria","provincia","localidad","nombre","direccion","cp","telefono","mail","web"]
    columnas_a_renombrar=["cod_localidad","id_provincia","id_departamento","categoría","provincia","localidad","nombre","domicilio","código postal","número de teléfono","mail","web"]
    a = a.drop(columns=[col for col in a if unidecode(col).lower() not in columnas_a_filtrar])
    b = b.drop(columns=[col for col in b if unidecode(col).lower() not in columnas_a_filtrar])
    c = c.drop(columns=[col for col in c if unidecode(col).lower() not in columnas_a_filtrar])

    # a.columns(columnas_a_renombrar)
    # b.columns(columnas_a_renombrar)
    # c.columns(columnas_a_renombrar)

    print(a.head)
    print(b.head)
    print(c.head)

    # df=pd.concat([a,b],sort=False)
    # df1=df
    # df2=pd.concat([df1,c],sort=False)
    # df_final=df2


    # print(df_final.head)

    # wb = Workbook()

    # ws =  wb.active
    # ws.title = "Test"

    # wb.save(filename = 'dataframe.xlsx')
    # book = load_workbook('dataframe.xlsx')
    # writer = pd.ExcelWriter('dataframe.xlsx','openpyxl')
    # writer.book = book
    # df_final.to_excel(writer, sheet_name = "Test",index=None)
    # writer.save()
    # writer.close()




    # merged = a.merge(b, on='title')
    # merged.to_csv("output.csv", index=False)